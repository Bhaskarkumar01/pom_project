# import xlrd

# path = r'C:\Users\KIIT\PycharmProjects\selenium_project\selenium_training\pom_project\files_\demo_testdata.xlsx'


# def read_excel():
#     workbook = xlrd.open_workbook(path)                     ## book obj
#     worksheet = workbook.sheet_by_name("Sheet1")            ## sheet_obj
#     rows = worksheet.get_rows()                             ## generator_obj
#     header = next(rows)

#     d = {}
#     for ele in rows:
#         d[ele[0].value] =  ele[1].value

#     return d


# import xlrd

# workbook = xlrd.open_workbook(path)
# sheet = workbook.sheet_by_index(0)



# def read_excel():
#     path = "testdata.xlsx"

#     workbook = load_workbook(path)
#     sheet = workbook.active

#     data = []

#     for row in sheet.iter_rows(values_only=True):
#         data.append(row)

#     return data



# -----------------------------------------------------------------------------


from openpyxl import load_workbook

path = r'C:\Users\KIIT\PycharmProjects\selenium_project\selenium_training\pom_project\files_\demo_testdata.xlsx'


def read_excel():
    workbook = load_workbook(path)        # workbook object
    worksheet = workbook["Sheet1"]        # sheet object

    rows = worksheet.iter_rows(values_only=True)

    header = next(rows)                   # skip header row

    d = {}
    for row in rows:
        d[row[0]] = row[1]                # column A → key, column B → value

    return d













































































